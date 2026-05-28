import openpyxl
import csv
import datetime

# Leer CSV
resultados = []
with open('c:/POS-Papeleria/Resultados.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        resultados.append(row)

# Cargar plantilla de Excel
wb = openpyxl.load_workbook('c:/POS-Papeleria/Tabla-de-resultados-de-las-pruebas-v2 (1).xlsx')
ws = wb.active

now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# --- PRIMERA TABLA: Recargas (filas 2-11) ---
recargas = [r for r in resultados if r['Tipo'] == 'Recarga']
for i, row in enumerate(recargas):
    excel_row = i + 2
    ws[f'E{excel_row}'] = now_str                      # FECHA/HORA
    ws[f'F{excel_row}'] = row['TransID']               # TRANS ID
    ws[f'G{excel_row}'] = row['Folio']                 # FOLIO
    ws[f'H{excel_row}'] = row['Status'].strip()        # ESTATUS

# --- SEGUNDA TABLA: Servicios (filas 14-19) ---
servicios = [r for r in resultados if r['Tipo'] == 'Servicio']
for i, row in enumerate(servicios):
    excel_row = i + 14
    ws[f'E{excel_row}'] = now_str                      # FECHA/HORA
    ws[f'F{excel_row}'] = row['TransID']               # TRANS ID
    ws[f'G{excel_row}'] = row['Folio']                 # FOLIO
    ws[f'H{excel_row}'] = row['Status'].strip()        # ESTATUS

# Guardar
wb.save('c:/POS-Papeleria/Tabla_Lista_Para_Subir.xlsx')
print("Excel generado correctamente con los nuevos resultados!")
